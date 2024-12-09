from tkinter import filedialog

from openpyxl import load_workbook  # import a class from openpyxl package
import os


# Paths

def get_valid_file_path(default_path, title, filetypes):
    """
    Validate the given file path. If it doesn't exist, prompt the user to select a valid file.

    Args:
        default_path (str): Default file path to check.
        title (str): Title for the file dialog box.
        filetypes (list): Allowed file types for the file dialog.

    Returns:
        str: A valid file path.
    """
    if not os.path.exists(default_path):
        print(f"File not found: {default_path}")
        default_path = filedialog.askopenfilename(title=title, filetypes=filetypes)
        if not default_path:  # User canceled file selection
            raise FileNotFoundError(f"{title} is required but no file was selected.")
    return default_path


path_root_folder = filedialog.askdirectory(
    title="Select a Folder",
    initialdir="/"  # Default directory (optional)
)
print(path_root_folder)

path_raw_report = get_valid_file_path(
    f'{path_root_folder}/resource/raw report.xlsx',
    title='Select RAW Report', filetypes=[("Excel files", "*.xlsx;*.xls")])

path_save = "C:/Users/mzrss/Downloads/raw_report_output.xlsx"
path_db = get_valid_file_path(f'{path_root_folder}/resource/db.xlsx',
                              title='Select 4G Database',
                              filetypes=[("Excel files", "*.xlsx;*.xls")])  # path of database should be latest

path_gis = get_valid_file_path(f'{path_root_folder}/resource/RS Spectra_SCFT_4G GIS.xlsx',
                               title='Select 4G gis', filetypes=[("Excel files", "*.xlsx;*.xls")])

sec1_azimuth = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC1/AZM.jpg',
    title='Select SEC 1 Azimuth', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec1_clutter = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC1/CL.jpg',
    title='Select SEC 1 Clutter', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])

sec1_mt = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC1/MT.jpg',
    title='Select SEC 1 Mechanical Tilt', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec1_et = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC1/ET.jpg',
    title='Select SEC 1 Electrical Tilt', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec1_label = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC1/LABEL.jpg',
    title='Select SEC 1 Label', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec2_azimuth = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC2/AZM.jpg',
    title='Select SEC 2 Azimuth', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec2_clutter = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC2/CL.jpg',
    title='Select SEC 2 Clutter', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec2_mt = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC2/MT.jpg',
    title='Select SEC 2 Mechanical Tilt', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec2_et = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC2/ET.jpg',
    title='Select SEC 2 Electrical Tilt', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec2_label = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC2/label.jpg',
    title='Select SEC 2 Label', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec3_azimuth = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC3/AZM.jpg',
    title='Select SEC 3 Azimuth', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec3_clutter = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC3/CL.jpg',
    title='Select SEC 3 Clutter', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec3_mt = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC3/MT.jpg',
    title='Select SEC 3 Mechanical Tilt', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec3_et = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC3/ET.jpg',
    title='Select SEC 3 Electrical Tilt', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec3_label = get_valid_file_path(
    f'{path_root_folder}/Audit photos/SEC3/lABEL.jpg',
    title='Select SEC 3 Label', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec1_tape = get_valid_file_path(
    f'{path_root_folder}/Audit photos/TOWER+ANTENNA/SEC1 AH.jpg',
    title='Select SEC 1 Antina Height', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec2_tape = get_valid_file_path(
    f'{path_root_folder}/Audit photos/TOWER+ANTENNA/SEC2 AH.jpg',
    title='Select SEC 2 Antina Height', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
sec3_tape = get_valid_file_path(
    f'{path_root_folder}/Audit photos/TOWER+ANTENNA/SEC3 AH.jpg',
    title='Select SEC 3 Antina Height', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])

tower = get_valid_file_path(
    f'{path_root_folder}/Audit photos/TOWER+ANTENNA/TP.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_0 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/0.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_30 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/30.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_60 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/60.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_90 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/90.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_120 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/120.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_150 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/150.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_180 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/180.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_210 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/210.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_240 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/240.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_270 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/270.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_300 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/300.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])
degree_330 = get_valid_file_path(
    f'{path_root_folder}/Audit photos/PANORAMIC/330.jpg',
    title='Select Tower Photo', filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")])

header_list = ['SITE', 'CELL', 'TECHNOLOGY', 'TOWER TYPE', 'TOWER HEIGHT', 'BUILDING HEIGHT', 'ANTINA HEIGHT',
               'PRE AZIMUTH', 'M TILT', 'E TILT', 'POST AZIMUTH', 'M TILT',
               'E TILT']  # heading values for audit details sheet

wb = load_workbook(path_raw_report)  # creating an object to raw report
telecom_sheet = wb['Telecom']  # assign telecom sheet from the raw report to telecom_sheet class
main_sheet = wb['Main Sheet']  # assign Main sheet  from the raw report to main_sheet class

VOLTE_SCFT_NEW_INT_sheet = wb[
    'VOLTE_SCFT_NEW_INT']  # assign VOLTE_SCFT_NEW_INT from the raw report to VOLTE_SCFT_NEW_INT_sheet class
wb.remove(wb['Clutter_Photos'])  # for deleting existing sheet because cell space not correct
wb.create_sheet('Clutter Photos', index=17)  # creating new clutter photos sheet
Clutter_Photos_sheet = wb['Clutter Photos']  # assign Clutter_Photos from the raw report to class
wb.create_sheet('Panoramic Photos', index=18)  # creating new sheet for panoramic photos
Panoramic_photos_sheet = wb['Panoramic Photos']  # create an object for panoramic photos sheet
wb.create_sheet("Audit Details", index=19)  # create a sheet for audit details
Audit_details_sheet = wb['Audit Details']  # create an object for audit details sheet
scft_sheet = wb["SCFT"]  # assign scft from raw report
SCFT_INT_sheet = wb['SCFT_INT']  # assign SCFT_INT from raw report
