from copy import copy
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import attributes as att
from openpyxl.drawing.image import Image

__gis = pd.read_excel(att.path_gis, sheet_name='4G')
__db = pd.read_excel(att.path_db, sheet_name='4G')


# function for copy cell format from telecom_sheet['D5']
def copy_cell_format(source_cell, target_cell):
    # Copy font
    target_cell.font = copy(source_cell.font)
    # Copy fill
    # target_cell.fill = copy(source_cell.fill)
    # Copy border
    target_cell.border = copy(source_cell.border)
    # Copy alignment
    target_cell.alignment = copy(source_cell.alignment)
    # Copy number format
    target_cell.number_format = source_cell.number_format
    # Copy protection
    target_cell.protection = copy(source_cell.protection)


# function for vlookup from db currently for district need to do dynamic
def vlookup(lookup_value):
    result = __db[__db['Site ID'] == lookup_value]
    if not result.empty:
        district = result.iloc[0, 80]  # Adjusting for zero-based index (81st column is index 80)
        return district
    else:
        district = None
        return district


# VOLTE_SCFT_NEW_INT**
# __recode_start___
def add_values_to_volte_scft(cells):
    sheet_name = att.VOLTE_SCFT_NEW_INT_sheet
    rows = {
        '5': 'N/A',
        '6': 100,
        # '7': random.uniform(20, 25),
        # '8': random.uniform(19, 22), manually entered in main sheet
        '9': 'N/A',
        '10': '0',
        '11': '0',
        '12': 'YES',
        '13': 'YES',
        '14': '0'
    }
    for cell in cells:
        for row, value in rows.items():
            sheet_name[f"{cell}{row}"].value = value
    # att.VOLTE_SCFT_NEW_INT_sheet[coloumn_name + '5'].value = 'N/A'
    # att.VOLTE_SCFT_NEW_INT_sheet[coloumn_name + '6'].value = 100
    # att.VOLTE_SCFT_NEW_INT_sheet[coloumn_name + '7'].value = random.uniform(20, 25)
    # att.VOLTE_SCFT_NEW_INT_sheet[coloumn_name + '8'].value = random.uniform(20, 25)
    # att.VOLTE_SCFT_NEW_INT_sheet[coloumn_name + '9'].value = 'N/A'
    # att.VOLTE_SCFT_NEW_INT_sheet[coloumn_name + '10'].value = '0'
    # att.VOLTE_SCFT_NEW_INT_sheet[coloumn_name + '11'].value = '0'
    # att.VOLTE_SCFT_NEW_INT_sheet[coloumn_name + '12'].value = 'YES'
    # att.VOLTE_SCFT_NEW_INT_sheet[coloumn_name + '13'].value = 'YES'
    # att.VOLTE_SCFT_NEW_INT_sheet[coloumn_name + '14'].value = '0'


# __recode_end__

# to add Pass value to particular range(min_row=3, max_row=14, min_col=7, max_col=9) of cells
def add_pass_to_volte_scft():
    value_to_set = 'Pass'
    for row in att.VOLTE_SCFT_NEW_INT_sheet.iter_rows(min_row=3, max_row=14, min_col=7, max_col=9):
        for cell in row:
            cell.value = value_to_set


# format the particular range(min_row=3, max_row=14, min_col=4, max_col=9) of cells from VOLTE_SCFT_NEW_INT_sheet['D3']
def format_to_volte_scft(source_cell_volte):
    for row in att.VOLTE_SCFT_NEW_INT_sheet.iter_rows(min_row=3, max_row=14, min_col=4, max_col=9):
        for cell in row:
            copy_cell_format(source_cell_volte, cell)


# format the clutter photos sheet such as merging heading bold
def sheet_cell_format(start_cell, end_cell, header_name, sheet_name):
    sheet_name.merge_cells(f'{start_cell}:{end_cell}')
    sheet_name[start_cell].value = header_name
    light_blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    sheet_name[start_cell].fill = light_blue_fill
    sheet_name[start_cell].font = Font(size=14, color='FFFFFF', bold=True)
    sheet_name[start_cell].alignment = Alignment(horizontal='center', vertical='center')
    thick_border = Border(
        left=Side(style='thick', color='000000'),  # Left border
        right=Side(style='thick', color='000000'),  # Right border
        top=Side(style='thick', color='000000'),  # Top border
        bottom=Side(style='thick', color='000000')  # Bottom border
    )
    for row in sheet_name[f'{start_cell}:{end_cell}']:
        for cell in row:
            cell.border = thick_border


# Adding images to clutter sheet
def sheet_add_image(image_path, cell, sheet_name):
    img = Image(image_path)
    img.width = 320
    img.height = 360
    sheet_name.add_image(img, cell)


# Adding heading values in Audit details sheet
def add_header_value(source_cell):  # an argument is coming format of source cell
    k = 1
    for i in att.header_list:
        j = 2
        att.Audit_details_sheet.cell(j, k).value = i  # adding list value into excel
        target_cell = att.Audit_details_sheet.cell(j, k)  # for format the cell to assign the target cell
        copy_cell_format(source_cell, target_cell)  # call function for format the header cell
        k = k + 1


# finding the tower type using function need recode and simplify with vlookup function
def tower_type(lookup_value):
    result = __gis[__gis['site'] == lookup_value]
    if not result.empty:
        tower = result.iloc[1, 17]  # Adjusting for zero-based index (81st column is index 80)
        return tower
    else:
        tower = None
        return tower


# finding tower height need to recode and simplify with vlookup function
def tower_height(lookup_value):
    result = __gis[__gis['site'] == lookup_value]
    if not result.empty:
        tower = result.iloc[1, 19]  # Adjusting for zero-based index (81st column is index 80)
        return tower
    else:
        tower = None
        return tower


#  finding building height if tower type is rtt
def tower_type_rtt(lookup_value):
    result = __gis[__gis['site'] == lookup_value]
    print(result)
    if not result.empty:
        tower = result.iloc[1, 18]
        return tower
    else:
        tower = None
        return tower


# finding pre azimuth need to recode
def azi_pre(lookup_value):
    result = __gis[__gis['SITE CELL'] == lookup_value]
    if not result.empty:
        tower = result.iloc[0, 21]
        return tower
    else:
        tower = None
        return tower


# finding post azimuth need to recode
def azi_post(lookup_value):
    result = __gis[__gis['SITE CELL'] == lookup_value]
    if not result.empty:
        tower = result.iloc[0, 25]
        return tower
    else:
        tower = None
        return tower
